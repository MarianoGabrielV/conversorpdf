import os
from openpyxl import load_workbook
from fpdf import FPDF
try:
    from docx2pdf import convert as docx_convert
except ImportError:
    docx_convert = None  # Si no está instalado, manejamos mejor el error

def convertir_word(ruta_archivo, carpeta_salida=None):
    if docx_convert is None:
        print("docx2pdf no está instalado o no disponible en este ejecutable.")
        return False

    try:
        if carpeta_salida:
            nombre_archivo = os.path.basename(ruta_archivo)
            ruta_temp = os.path.join(carpeta_salida, nombre_archivo)
            os.system(f'copy "{ruta_archivo}" "{ruta_temp}"')
            docx_convert(ruta_temp)
            os.remove(ruta_temp)
        else:
            docx_convert(ruta_archivo)
        return True
    except Exception as e:
        print("⚠️ Microsoft Word no encontrado o problema al convertir Word a PDF.")
        print(f"Detalle del error: {e}")
        return False

def convertir_excel(ruta_archivo, carpeta_salida=None):
    try:
        ruta_archivo = os.path.abspath(ruta_archivo)
        print(f"Intentando abrir archivo: {ruta_archivo}")

        wb = load_workbook(ruta_archivo)
        ws = wb.active

        if carpeta_salida:
            nombre_salida = os.path.splitext(os.path.basename(ruta_archivo))[0] + ".pdf"
            ruta_pdf = os.path.join(carpeta_salida, nombre_salida)
        else:
            ruta_pdf = ruta_archivo.replace('.xlsx', '.pdf')

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=10)

        num_cols = ws.max_column
        col_width = pdf.w / (num_cols + 1)

        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                pdf.set_fill_color(200, 200, 200)  # Encabezado gris
                pdf.set_font("Arial", 'B', 11)
                fill = True
            else:
                pdf.set_fill_color(255, 255, 255)  # Blanco normal
                pdf.set_font("Arial", '', 10)
                fill = False

            for cell in row:
                contenido = str(cell) if cell is not None else ""
                pdf.cell(col_width, 10, contenido, border=1, fill=fill)
            pdf.ln(10)

        pdf.output(ruta_pdf)

        return True
    except Exception as e:
        print(f"Error al convertir Excel: {e}")
        return False
